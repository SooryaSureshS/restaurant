# -*- coding: utf-8 -*-


from odoo import models, fields, _, api
from datetime import timedelta, datetime, date

from odoo.exceptions import ValidationError
import io
import base64
import os.path
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart,
)
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter


class OperationalReportWizard(models.TransientModel):
    _name = "operational.report.wizard"
    
    report_summary_file = fields.Binary('Generated Report')
    file_name = fields.Char('File Name')
    # order_no = fields.Char(string="Order No")
    from_date = fields.Date(string="From Date", default=datetime.today())
    to_date = fields.Date(string="To Date", default=datetime.today())
    # order_date = fields.Date(string="Order Date", default=datetime.today())
    # picking_time = fields.Date(string="Picking Time", default=datetime.today())

    @api.constrains('from_date', 'to_date')
    def _date_check(self):
        today = date.today()
        if not self.from_date:
            raise ValidationError(_('Select From Date'))
        if not self.to_date:
            raise ValidationError(_('Select To Date'))
        if self.from_date and self.to_date:
            if self.from_date > self.to_date:
                raise ValidationError(_('From date should be less than To Date'))
            if self.from_date > today:
                raise ValidationError(_('Future Dates are not allowed'))
            if self.to_date > today:
                raise ValidationError(_('Future Dates are not allowed'))
    def print_report(self):
        return self.env.ref('operational_report.operational_xlsx_report').report_action(self, data=self.read([])[0])


    def print_graph_report(self):
        print("graph ...........report")
        if self.from_date and self.to_date:
            wb = Workbook()
            ws = wb.active
            row_custom=5
            month_row=4
            bar_row = 4
            fp = io.BytesIO()
            charts_row=1
            last_list=[]
            month = []
            total_sa=[]

            sale_obj = self.env['sale.order'].search([('date_order','>','%s 00:00:00' % self.from_date),
                                                       ('date_order','<','%s 23:59:59' % self.to_date)])

            last_list.append(tuple(('Date', 'No. of SO'),))
            from_date_limit = int(self.from_date.strftime('%m'))-1
            to_date_limit = int(self.to_date.strftime('%m'))
            summary_list=[]
            while from_date_limit < to_date_limit:
                list_p = []
                date_now = self.from_date + relativedelta(months=from_date_limit)  # get month from date
                date_string = str(date_now.strftime('%h')) + ' ' + str(date_now.year)
                month.append(date_string)
                from_date_limit += 1
                total_sale = len(sale_obj.filtered(lambda x:date_now.strftime('%Y-%m')==x.date_order.strftime('%Y-%m')))
                row=row_custom
                total_sa.append(total_sale)

                list_p.append(date_string)
                list_p.append(total_sale)
                last_list.append(tuple(list_p))

            print('last_list',last_list)
            for row in last_list:
                ws.append(row)

            data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(last_list))
            titles = Reference(ws, min_col=1, min_row=2, max_row=len(last_list))

            chart1 = BarChart()
            chart1.title = "Bar Chart"

            chart1.add_data(data=data, titles_from_data=True)
            chart1.set_categories(titles)
            chart1.x_axis.delete = False
            chart1.y_axis.delete = False
            ws.add_chart(chart1, "D10")

            wb.save(fp)
            excel_file = base64.encodebytes(fp.getvalue())
            self.report_summary_file = excel_file
            self.file_name = 'Graph Report xlsx'
            fp.close()

            return {
                'view_mode': 'form',
                'res_id': self.id,
                'res_model': 'operational.report.wizard',
                'view_type': 'form',
                'type': 'ir.actions.act_window',
                'context': self.env.context,
                'target': 'new'}

#         return self.env.ref('operational_report.graph_operational_xlsx_report').report_action(self, data=self.read([])[0])


